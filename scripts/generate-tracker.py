"""
Generate VirtualOffice AI — Product Tracker.xlsx
Run: python3 scripts/generate-tracker.py
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

# ─── Colour palette ────────────────────────────────────────────────────────────
C_DONE      = "1E4D2B"   # dark green  text on done rows
C_DONE_BG   = "D6ECD9"
C_ACTIVE_BG = "D0E8F7"
C_ACTIVE    = "0D2F4A"
C_PEND_BG   = "FFFBEC"
C_PEND      = "4A3A00"
C_DEFERRED  = "F0F0F0"
C_DEFER_TXT = "888888"

C_HDR_PHASE = "1F3B5C"   # phase header row
C_HDR_COL   = "2C5282"   # column header row
C_WHITE     = "FFFFFF"
C_BORDER    = "BBBBBB"

STATUS_STYLE = {
    "Done":      (C_DONE_BG,   C_DONE,    "✅ Done"),
    "Active":    (C_ACTIVE_BG, C_ACTIVE,  "▶ Active"),
    "Pending":   (C_PEND_BG,   C_PEND,    "○ Pending"),
    "Deferred":  (C_DEFERRED,  C_DEFER_TXT,"⏸ Deferred"),
}

thin = Side(style="thin", color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def row_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def phase_font():
    return Font(bold=True, color=C_WHITE, size=11)

def col_font():
    return Font(bold=True, color=C_WHITE, size=10)

def body_font(color="000000", bold=False, size=10):
    return Font(color=color, bold=bold, size=size)

def wrap(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# ─── Data ──────────────────────────────────────────────────────────────────────
# Columns: Phase | Sprint | Feature | Sub-feature | Status | Owner | Notes
ROWS = [
    # ── PHASE 1 ──────────────────────────────────────────────────────────────
    ("PHASE", "Phase 1 — Foundation & Staging", "", "", "", "", ""),

    ("P1", "S1 — Core Architecture", "TypeScript Monorepo", "pnpm workspaces + Turborepo, 19 packages", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "TypeScript Monorepo", "tsconfig references + noUncheckedIndexedAccess", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "TypeScript Monorepo", "ESLint + Prettier + Husky pre-commit", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Multi-tenant Data Layer", "TenantContext explicit value type (ADR 007)", "Done", "Engineering", "No AsyncLocalStorage"),
    ("P1", "S1 — Core Architecture", "Multi-tenant Data Layer", "withTenant pattern — RLS via app.tenant_id", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Multi-tenant Data Layer", "Two-role Postgres: voai_admin (migrations) + voai_app (runtime)", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Multi-tenant Data Layer", "Row-Level Security FORCE on all tenant tables", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Multi-tenant Data Layer", "pgvector extension for brain embeddings", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Local Dev Stack", "Docker Compose: Postgres + pgvector + Redis + Neo4j", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Local Dev Stack", "DB migration runner (node-pg-migrate)", "Done", "Engineering", ""),
    ("P1", "S1 — Core Architecture", "Local Dev Stack", "Turborepo dev task with watch mode", "Done", "Engineering", ""),

    ("P1", "S2 — Core Services", "identity-and-tenancy", "Tenant creation + scoped DB bootstrap", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "identity-and-tenancy", "User creation under tenant", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "identity-and-tenancy", "Session creation + TenantContext assembly", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "brain", "Memory storage (vector embeddings via pgvector)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "brain", "Semantic search: retrieve relevant prior context", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "brain", "Neo4j graph: entity + relationship extraction", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "brain", "Brain context injected into agent prompts across sessions", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "ledger", "decisions table (type, summary, state)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "ledger", "actions table (assigned_to, state, due_at)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "ledger", "conflicts table (type, severity, resolution_state)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "ledger", "CRUD endpoints for all three journal types", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "routing", "AnthropicProvider implementing LlmProvider interface", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "routing", "RoutingService: single seam for all LLM calls", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "routing", "Routing decision structured logging", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Single-agent contribution (generateContribution)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Roster dispatch — fan-out to all personas in parallel", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Persona: Rajan Mehta (CEO / Strategic Advisor)", "Done", "Product", "Sarah Chen in code"),
    ("P1", "S2 — Core Services", "agent-runtime", "Persona: Priya Reddy (CMO)", "Done", "Product", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Persona: Marcus Webb (Devil's Advocate)", "Done", "Product", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Brain context injection per-turn", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "agent-runtime", "Teammate awareness (no hallucinated names)", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "meeting", "Session create / get / list endpoints", "Done", "Engineering", ""),
    ("P1", "S2 — Core Services", "Demo Web UI", "Thin React UI: send message, see 3 persona responses", "Done", "Product", ""),

    ("P1", "S3 — Wave 3 Services", "performance", "6-signal schema: factual_grounding, peer_agreement, expert_agreement, founder_action, outcome, pushback", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "performance", "POST /signal — record a signal for a persona", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "performance", "GET /scores/:personaId — aggregate by window (default 30d)", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "performance", "GET /weekly — ranked weekly leaderboard", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "notification", "notification_preferences table + RLS", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "notification", "GET /briefing — Claude Haiku morning briefing (24h activity summary)", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "notification", "GET/PUT /preferences — per-tenant briefing settings", "Done", "Engineering", ""),
    ("P1", "S3 — Wave 3 Services", "notification", "POST /alert — log high-risk or conflict alert", "Done", "Engineering", ""),

    ("P1", "S4 — AWS Staging", "Infrastructure", "ECR repository + IAM policy (AmazonEC2ContainerRegistryPowerUser)", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Infrastructure", "ECS Fargate cluster + task definition (linux/amd64)", "Done", "Infra", "ADR 012: ECS over App Runner"),
    ("P1", "S4 — AWS Staging", "Infrastructure", "Application Load Balancer + target group + /healthz health check", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Infrastructure", "RDS Postgres 16.9 (ap-south-1)", "Done", "Infra", "16.4 not available in region"),
    ("P1", "S4 — AWS Staging", "Infrastructure", "ElastiCache Redis", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Infrastructure", "S3 object store bucket (voai-staging-objects)", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Infrastructure", "Secrets Manager: DATABASE_URL, MIGRATIONS_DATABASE_URL, REDIS_URL, NEO4J, ANTHROPIC_API_KEY", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Infrastructure", "VPC private subnets, security groups", "Done", "Infra", ""),
    ("P1", "S4 — AWS Staging", "Database", "voai_admin + voai_app roles created on RDS", "Done", "Engineering", ""),
    ("P1", "S4 — AWS Staging", "Database", "All 7 migrations applied (0000–0007)", "Done", "Engineering", "Includes performance + notification"),
    ("P1", "S4 — AWS Staging", "Database", "RLS policies active on staging RDS", "Done", "Engineering", ""),
    ("P1", "S4 — AWS Staging", "CI/CD", "GitHub Actions deploy-staging.yml", "Done", "Engineering", "Manual push currently; no GitHub remote yet"),
    ("P1", "S4 — AWS Staging", "CI/CD", "GitHub remote + repo secrets (AWS keys)", "Done", "Engineering", "manuani/mysquad; CI/Deploy both green"),

    # ── PHASE 2 ──────────────────────────────────────────────────────────────
    ("PHASE", "Phase 2 — Live Meeting Intelligence", "", "", "", "", ""),

    ("P2", "S5 — Response Gating", "agent-runtime", "checkShouldRespond() — Haiku gate, max 80 tokens, returns JSON {shouldRespond, relevanceScore, reason}", "Done", "Engineering", "Deployed sprint5 tag"),
    ("P2", "S5 — Response Gating", "agent-runtime", "generateRosterContributions — skip personas below 0.4 threshold", "Done", "Engineering", ""),
    ("P2", "S5 — Response Gating", "agent-runtime", "Fail-open: JSON parse error → shouldRespond=true", "Done", "Engineering", ""),
    ("P2", "S5 — Response Gating", "agent-runtime", "skipGate option for callers that need all-or-nothing dispatch", "Done", "Engineering", ""),
    ("P2", "S5 — Response Gating", "agent-runtime", "Response includes skipped + gateResult per entry", "Done", "Engineering", ""),
    ("P2", "S5 — Response Gating", "Tests", "Gate unit tests: skip on low score, fail-open on bad JSON, gateResult exposed", "Done", "Engineering", "22 tests total"),

    ("P2", "S6 — Proactive Observer Loop", "agent-runtime", "observeSkippedPersonas(): re-score skipped personas with 0.65 threshold after roster completes", "Done", "Engineering", "Deployed sprint6"),
    ("P2", "S6 — Proactive Observer Loop", "agent-runtime", "Publish raise-hand event on EventBus (non-blocking, fire-and-forget)", "Done", "Engineering", ""),
    ("P2", "S6 — Proactive Observer Loop", "meeting", "SseManager: track SSE connections per sessionId, fan-out on event", "Done", "Engineering", ""),
    ("P2", "S6 — Proactive Observer Loop", "meeting", "SSE endpoint: GET /v1/meeting/sessions/:id/events (query-param auth for EventSource)", "Done", "Engineering", ""),
    ("P2", "S6 — Proactive Observer Loop", "meeting", "Subscribe to raise-hand EventBus events → SseManager.emit()", "Done", "Engineering", ""),
    ("P2", "S6 — Proactive Observer Loop", "Demo UI", "EventSource SSE connection on session start", "Done", "Product", ""),
    ("P2", "S6 — Proactive Observer Loop", "Demo UI", "Yellow hand-raise bar with per-persona chips (name + confidence %)", "Done", "Product", ""),
    ("P2", "S6 — Proactive Observer Loop", "Demo UI", "Click chip → fetch solo contribution; ✕ chip → dismiss", "Done", "Product", ""),

    ("P2", "S7 — LLM Arbiter", "agent-runtime", "Arbiter: when ≥2 personas pass gate simultaneously, rank by (role relevance, silence penalty, perf score)", "Done", "Engineering", "compositeScore = relevance×0.65 + silence×0.35"),
    ("P2", "S7 — LLM Arbiter", "agent-runtime", "Ordered contribution queue — personas speak in ranked order, not parallel", "Done", "Engineering", "generateOrderedContributions() with cumulative context"),
    ("P2", "S7 — LLM Arbiter", "agent-runtime", "Configurable max-speakers-per-turn limit", "Done", "Engineering", "DEFAULT_MAX_SPEAKERS=2"),
    ("P2", "S7 — LLM Arbiter", "performance", "Silence penalty integration: use weekly score to boost quiet personas", "Done", "Engineering", "silenceScore = min(minutesSilent/30, 1.0)"),

    ("P2", "S8 — Voice Pipeline", "Infrastructure", "LiveKit Cloud account + project", "Pending", "Infra", "Requires manual account creation + LIVEKIT_* env vars"),
    ("P2", "S8 — Voice Pipeline", "Infrastructure", "LiveKit room-per-session provisioning", "Pending", "Infra", "Token per session issued by meeting service"),
    ("P2", "S8 — Voice Pipeline", "Infrastructure", "Deepgram streaming STT account + API key in Secrets Manager", "Pending", "Infra", "SDK wired; set DEEPGRAM_API_KEY to activate"),
    ("P2", "S8 — Voice Pipeline", "Infrastructure", "ElevenLabs TTS account + voice IDs per persona", "Pending", "Infra", "SDK wired; set ELEVENLABS_API_KEY to activate"),
    ("P2", "S8 — Voice Pipeline", "media-coordinator", "Audio pipeline coordinator: receive STT transcript chunks", "Done", "Engineering", "apps/media-coordinator Express service on port 3001"),
    ("P2", "S8 — Voice Pipeline", "media-coordinator", "Route transcript chunks to agent-runtime", "Done", "Engineering", "pipeline.ts: STT → /v1/agent-runtime/contributions/roster"),
    ("P2", "S8 — Voice Pipeline", "media-coordinator", "Receive agent text response → send to ElevenLabs TTS", "Done", "Engineering", "tts.ts createTtsClient; graceful no-op when key absent"),
    ("P2", "S8 — Voice Pipeline", "media-coordinator", "Play TTS audio back into LiveKit room track", "Pending", "Engineering", "Needs LiveKit Node SDK room join — deferred to native app"),
    ("P2", "S8 — Voice Pipeline", "meeting", "LiveKit token generation endpoint (founder joins audio room)", "Done", "Engineering", "POST /sessions/:id/voice-token → AccessToken JWT"),
    ("P2", "S8 — Voice Pipeline", "Demo UI", "MediaRecorder audio → media-coordinator (no LiveKit browser SDK)", "Done", "Product", "250ms chunks; polls /status every 800ms"),
    ("P2", "S8 — Voice Pipeline", "Demo UI", "Real-time transcript display (live chunk preview)", "Done", "Product", "transcript-chunk div updated on each 800ms poll"),
    ("P2", "S8 — Voice Pipeline", "Demo UI", "Persona voice labels (which voice is speaking)", "Done", "Product", "voice-personas.ts maps name → ElevenLabs voice ID"),

    # ── PHASE 3 ──────────────────────────────────────────────────────────────
    ("PHASE", "Phase 3 — Expert Network & Marketplace", "", "", "", "", ""),

    ("P3", "S9 — Expert Profiles & Matching", "marketplace", "Expert schema: domain tags, availability windows, rate card", "Done", "Engineering", "migration 0008; expert_profiles + expert_domain_tags + expert_availability"),
    ("P3", "S9 — Expert Profiles & Matching", "marketplace", "Expert onboarding flow (API)", "Done", "Product", "POST /experts + PATCH /experts/:id + POST /experts/:id/tags"),
    ("P3", "S9 — Expert Profiles & Matching", "marketplace", "Expert matching endpoint: given topic → ranked expert list", "Done", "Engineering", "POST /match; token-overlap scoring; Phase 4 upgrades to vector similarity"),
    ("P3", "S9 — Expert Profiles & Matching", "marketplace", "Escalation events: persona → real expert suggestion", "Done", "Engineering", "POST /escalations + PATCH /escalations/:id (accept/dismiss)"),
    ("P3", "S9 — Expert Profiles & Matching", "agent-runtime", "Persona can trigger 'escalate to real expert' action", "Done", "Engineering", "POST /v1/agent-runtime/escalate; matchExperts + escalation.triggered event; sprint13"),
    ("P3", "S9 — Expert Profiles & Matching", "agent-runtime", "Escalation event surfaced to founder in UI", "Done", "Product", "Expert match chips in demo UI after each roster response; sprint14"),
    ("P3", "S9 — Expert Profiles & Matching", "brain", "Expert domain knowledge indexed in Neo4j graph", "Done", "Engineering", "graph.ts: (Expert)-[:HAS_DOMAIN]->(Domain); graphMatchExperts; 7 tests; sprint14"),

    ("P3", "S10 — Expert Session Scheduling", "scheduler", "scheduler app: cron-style job runner", "Done", "Engineering", "apps/scheduler; CronRunner; parseCron; shouldRun; 13 tests"),
    ("P3", "S10 — Expert Session Scheduling", "scheduler", "Morning briefing cron job (08:00 UTC daily)", "Done", "Engineering", "calls /internal/scheduler/morning-briefing on api-server"),
    ("P3", "S10 — Expert Session Scheduling", "marketplace", "Expert availability slots API", "Done", "Engineering", "GET /experts/:id/slots?date=YYYY-MM-DD → 30-min slot list"),
    ("P3", "S10 — Expert Session Scheduling", "marketplace", "Expert session booking (Cal.com integration)", "Done", "Engineering", "POST /experts/:id/book; Cal.com when CALCOM_API_KEY set"),
    ("P3", "S10 — Expert Session Scheduling", "meeting", "Expert-join flow: invite expert to LiveKit room", "Done", "Engineering", "POST /sessions/:id/expert-join-token; expert- identity; 2 tests; sprint15"),
    ("P3", "S10 — Expert Session Scheduling", "Demo UI", "Expert availability picker + booking confirmation", "Done", "Product", "Booking modal on escalation chip click; slot list + confirm; sprint15"),

    ("P3", "S11 — Usage Metering & Billing", "marketplace-metering", "Token count per session stored in metering table", "Done", "Engineering", "POST /events; llm_tokens + ai_roster_call event types"),
    ("P3", "S11 — Usage Metering & Billing", "marketplace-metering", "Expert-minute billing events", "Done", "Engineering", "POST /events with eventType=expert_minutes"),
    ("P3", "S11 — Usage Metering & Billing", "marketplace-metering", "Monthly usage summary per tenant", "Done", "Engineering", "GET /usage?from=&to= → UsageSummary; monthly_usage_rollup table"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Stripe subscription tiers (starter/growth/enterprise)", "Done", "Engineering", "POST /billing/subscribe; stubs when STRIPE_SECRET_KEY absent"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Stripe charge per expert session", "Done", "Engineering", "POST /billing/expert-charge; invoice items"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Expert payout (Stripe Connect)", "Pending", "Engineering", "Requires Stripe Connect onboarding — deferred to Phase 4"),
    ("P3", "S11 — Usage Metering & Billing", "routing", "Token routing: track per-tenant LLM spend via RoutingService", "Done", "Engineering", "RoutingService onUsage callback → recordMeteringEvent llm_tokens; sprint13"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Stripe webhook handler (subscription lifecycle)", "Done", "Engineering", "POST /billing/webhook; subscription.created/updated/deleted → update tenant plan; sprint16"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Stripe Checkout session endpoint", "Done", "Engineering", "POST /billing/checkout → checkoutUrl; stub when STRIPE_SECRET_KEY absent; sprint20"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Plan entitlement limits + quota enforcement", "Done", "Engineering", "GET /entitlement?dim=; starter 100 calls, growth 1000; sprint17"),
    ("P3", "S11 — Usage Metering & Billing", "Payments", "Expert payout (Stripe Connect)", "Pending", "Engineering", "Requires Stripe Connect onboarding — deferred to Phase 4"),

    ("P3", "S12 — Admin Console", "admin-console-api", "Tenant provisioning endpoint — POST /tenants with x-admin-key auth", "Done", "Engineering", "sprint12"),
    ("P3", "S12 — Admin Console", "admin-console-api", "User management: invite, role assignment, deactivation", "Deferred", "Engineering", ""),
    ("P3", "S12 — Admin Console", "admin-console-api", "Usage dashboard — GET /tenants/:id/usage with metering breakdown", "Done", "Engineering", "sprint12"),
    ("P3", "S12 — Admin Console", "admin-console-api", "Seat-based subscription tier enforcement in routing layer", "Done", "Engineering", "checkEntitlement() seats dimension; sprint17"),
    ("P3", "S12 — Admin Console", "Admin UI", "Internal admin web app (separate from founder-facing demo UI)", "Deferred", "Product", ""),

    ("P3", "S13 — Credentials & Voice", "Secrets Manager", "LiveKit URL + API key + secret stored in Secrets Manager", "Done", "Infra", "ECS task definition :10"),
    ("P3", "S13 — Credentials & Voice", "Secrets Manager", "Deepgram API key stored in Secrets Manager", "Done", "Infra", "ECS task definition :11"),
    ("P3", "S13 — Credentials & Voice", "Secrets Manager", "ElevenLabs API key stored in Secrets Manager", "Done", "Infra", "ECS task definition :12"),
    ("P3", "S13 — Credentials & Voice", "Secrets Manager", "Voice persona IDs as plain ECS env vars", "Done", "Infra", "VOICE_ID_SARAH/PRIYA/MARCUS; ECS task definition :13"),
    ("P3", "S13 — Credentials & Voice", "Secrets Manager", "Stripe webhook secret stored in Secrets Manager", "Done", "Infra", "ECS task definition :14"),

    ("P3", "S14 — CI/CD & Migrations", "CI/CD", "GitHub repo (manuani/mysquad) + AWS secrets configured", "Done", "Engineering", "CI green; full build→typecheck→lint→test→deploy pipeline"),
    ("P3", "S14 — CI/CD & Migrations", "CI/CD", "Prettier format check in CI", "Done", "Engineering", "All files formatted; CI format:check passes"),
    ("P3", "S14 — CI/CD & Migrations", "Database", "Migration runner (packages/db/src/migrate.ts)", "Done", "Engineering", "schema_migrations tracking table; idempotent; sprint19"),
    ("P3", "S14 — CI/CD & Migrations", "Database", "Auto-run migrations in deploy-staging.yml before ECS deploy", "Done", "Engineering", "run-migrations job fetches URL from Secrets Manager; sprint19"),
    ("P3", "S14 — CI/CD & Migrations", "Database", "migrations 0009–0011 applied (metering, tenant_plan, stripe_customer_id)", "Done", "Engineering", ""),

    ("P3", "S15 — Demo UI Hardening", "Demo UI", "Plan tier badge in meeting room header", "Done", "Product", "Fetches /v1/metering/entitlement on init; sprint18"),
    ("P3", "S15 — Demo UI Hardening", "Demo UI", "Quota warning at 80% of monthly limit", "Done", "Product", "Amber badge showing current/limit; sprint18"),

    # ── PHASE 4 DEFERRED ─────────────────────────────────────────────────────
    ("PHASE", "Phase 4 — Scale & Production Hardening (Deferred)", "", "", "", "", ""),

    ("P4", "S13 — CI / DevOps", "CI/CD", "GitHub remote: create repo, push code, configure branch protection", "Done", "Engineering", "manuani/mysquad; AWS secrets configured; CI green"),
    ("P4", "S13 — CI / DevOps", "CI/CD", "GitHub Actions auto-deploy on push to main", "Done", "Engineering", "deploy-staging.yml; build→ECR→migrations→ECS forced deploy"),
    ("P4", "S13 — CI / DevOps", "CI/CD", "Terraform remote state — S3 backend migration (1 command; bucket exists)", "Deferred", "Infra", ""),
    ("P4", "S13 — CI / DevOps", "CI/CD", "Staging smoke test job in CI pipeline", "Deferred", "Engineering", ""),
    ("P4", "S13 — CI / DevOps", "IAM", "Narrow IAM policy (scoped policy replacing broad managed policies)", "Deferred", "Infra", "Before co-founder / contractor access"),

    ("P4", "S14 — Production Environment", "Infra", "prod VPC + separate RDS Multi-AZ instance", "Deferred", "Infra", ""),
    ("P4", "S14 — Production Environment", "Infra", "ECS desired-count 2 + ALB health-check tuning", "Deferred", "Infra", ""),
    ("P4", "S14 — Production Environment", "Infra", "Custom domain + ACM cert for ALB (HTTPS)", "Deferred", "Infra", ""),
    ("P4", "S14 — Production Environment", "Infra", "CloudFront CDN in front of ALB", "Deferred", "Infra", ""),
    ("P4", "S14 — Production Environment", "Infra", "WAF rules on ALB", "Deferred", "Infra", ""),

    ("P4", "S15 — Observability", "Monitoring", "CloudWatch dashboards: ECS CPU/mem, RDS connections, ALB latency", "Deferred", "Infra", ""),
    ("P4", "S15 — Observability", "Monitoring", "PagerDuty / SNS alerts on 5xx rate, task failures", "Deferred", "Infra", ""),
    ("P4", "S15 — Observability", "Monitoring", "Structured log search (CloudWatch Insights or Datadog)", "Deferred", "Infra", ""),
    ("P4", "S15 — Observability", "Monitoring", "Distributed trace IDs through routing + agent calls", "Deferred", "Engineering", ""),

    ("P4", "S16 — Multi-provider LLM Routing", "routing", "Four-tier LLM classification: Advanced / High / Good / OpenSource", "Deferred", "Engineering", "ADR — subscription-tier-driven routing"),
    ("P4", "S16 — Multi-provider LLM Routing", "routing", "OpenAI provider implementation", "Deferred", "Engineering", ""),
    ("P4", "S16 — Multi-provider LLM Routing", "routing", "Bedrock provider implementation", "Deferred", "Engineering", ""),
    ("P4", "S16 — Multi-provider LLM Routing", "routing", "Failover: if primary provider errors, retry on next tier", "Deferred", "Engineering", ""),
    ("P4", "S16 — Multi-provider LLM Routing", "routing", "Cost-per-call tracking per provider per tenant", "Deferred", "Engineering", ""),

    ("P4", "S17 — Background Workers", "worker", "worker app: BullMQ job queue over Redis (apps/worker stub)", "Deferred", "Engineering", ""),
    ("P4", "S17 — Background Workers", "worker", "Async brain indexing (don't block meeting response)", "Deferred", "Engineering", ""),
    ("P4", "S17 — Background Workers", "worker", "Async Neo4j graph updates", "Deferred", "Engineering", ""),
    ("P4", "S17 — Background Workers", "worker", "Dead-letter queue + retry with exponential backoff", "Deferred", "Engineering", ""),

    ("P4", "S18 — Security & Compliance", "Security", "HTTPS enforcement (redirect HTTP → HTTPS at ALB)", "Deferred", "Infra", ""),
    ("P4", "S18 — Security & Compliance", "Security", "Secrets rotation (Secrets Manager auto-rotate)", "Deferred", "Infra", ""),
    ("P4", "S18 — Security & Compliance", "Security", "Pen test / OWASP audit before public launch", "Deferred", "Engineering", ""),
    ("P4", "S18 — Security & Compliance", "Security", "GDPR data deletion endpoint (tenant offboarding)", "Deferred", "Engineering", ""),
    ("P4", "S18 — Security & Compliance", "Security", "Audit log: every data write recorded with actor + timestamp", "Deferred", "Engineering", ""),
]

# ─── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Tracker"

# ── Summary sheet ──────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Summary", 0)
ws.title = "All Tasks"

# ─── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [6, 28, 28, 58, 12, 12, 36]
COL_LABELS = ["Phase", "Sprint", "Feature / Service", "Sub-feature / Task", "Status", "Owner", "Notes"]

for i, (w, label) in enumerate(zip(COL_WIDTHS, COL_LABELS), 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ─── Column header row ─────────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# Row 1: title bar
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = f"VirtualOffice AI — Product Tracker   (generated {date.today()})"
title_cell.font = Font(bold=True, color=C_WHITE, size=13)
title_cell.fill = hdr_fill(C_HDR_PHASE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Row 2: column headers
for i, label in enumerate(COL_LABELS, 1):
    cell = ws.cell(row=2, column=i, value=label)
    cell.font = col_font()
    cell.fill = hdr_fill(C_HDR_COL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[2].height = 20

# ─── Data rows ─────────────────────────────────────────────────────────────────
ROW_OFFSET = 3
counts = {"Done": 0, "Active": 0, "Pending": 0, "Deferred": 0}

for r_idx, row_data in enumerate(ROWS, ROW_OFFSET):
    phase_tag, sprint, feature, subtask, status, owner, notes = row_data

    if phase_tag == "PHASE":
        # Phase separator row
        ws.merge_cells(f"A{r_idx}:G{r_idx}")
        cell = ws.cell(row=r_idx, column=1, value=sprint)
        cell.font = Font(bold=True, color=C_WHITE, size=11)
        cell.fill = hdr_fill(C_HDR_PHASE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = border
        ws.row_dimensions[r_idx].height = 22
        continue

    # Normal data row
    bg, fg, status_label = STATUS_STYLE.get(status, (C_PEND_BG, C_PEND, status))
    fill = row_fill(bg)

    vals = [phase_tag, sprint, feature, subtask, status_label, owner, notes]
    for c_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = body_font(color=fg, bold=(c_idx == 4))
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[r_idx].height = 30
    if status in counts:
        counts[status] += 1

# ─── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 40

ws_sum.merge_cells("A1:C1")
s = ws_sum["A1"]
s.value = "VirtualOffice AI — Tracker Summary"
s.font = Font(bold=True, color=C_WHITE, size=13)
s.fill = hdr_fill(C_HDR_PHASE)
s.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 28

headers = ["Status", "Count", "Description"]
for i, h in enumerate(headers, 1):
    c = ws_sum.cell(row=2, column=i, value=h)
    c.font = col_font()
    c.fill = hdr_fill(C_HDR_COL)
    c.alignment = Alignment(horizontal="center")
    c.border = border

summary_data = [
    ("✅ Done",     counts["Done"],     "Shipped and deployed to staging"),
    ("▶ Active",   counts["Active"],   "Currently in progress"),
    ("○ Pending",  counts["Pending"],  "Defined, not yet started"),
    ("⏸ Deferred", counts["Deferred"], "Phase 4 — after Phase 3 ships"),
]
total = sum(counts.values())

for i, (st, cnt, desc) in enumerate(summary_data, 3):
    bg, fg, _ = STATUS_STYLE.get(st.split()[-1], (C_PEND_BG, C_PEND, ""))
    # match by last word
    matched = None
    for k in STATUS_STYLE:
        if k in st:
            matched = STATUS_STYLE[k]
            break
    bg_c, fg_c = (matched[0], matched[1]) if matched else (C_PEND_BG, C_PEND)

    ws_sum.cell(row=i, column=1, value=st).font = body_font(color=fg_c, bold=True)
    ws_sum.cell(row=i, column=2, value=cnt).font = body_font(color=fg_c, bold=True)
    ws_sum.cell(row=i, column=3, value=desc).font = body_font(color=fg_c)
    for col in range(1, 4):
        ws_sum.cell(row=i, column=col).fill = row_fill(bg_c)
        ws_sum.cell(row=i, column=col).border = border
        ws_sum.cell(row=i, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    ws_sum.row_dimensions[i].height = 22

# Total row
ws_sum.cell(row=7, column=1, value="TOTAL").font = Font(bold=True, size=11)
ws_sum.cell(row=7, column=2, value=total).font = Font(bold=True, size=11)
ws_sum.cell(row=7, column=3, value="All tracked tasks").font = Font(size=10)
for col in range(1, 4):
    ws_sum.cell(row=7, column=col).border = border
    ws_sum.cell(row=7, column=col).alignment = Alignment(vertical="top")

# Completion % row
pct = round(counts["Done"] / total * 100) if total else 0
ws_sum.cell(row=8, column=1, value="Completion").font = Font(bold=True)
ws_sum.cell(row=8, column=2, value=f"{pct}%").font = Font(bold=True, color=C_DONE)
ws_sum.cell(row=8, column=2).fill = row_fill(C_DONE_BG)
for col in range(1, 4):
    ws_sum.cell(row=8, column=col).border = border

# Notes section
ws_sum.cell(row=10, column=1, value="Immediate Blockers").font = Font(bold=True, color="AA0000")
ws_sum.cell(row=11, column=1, value="GitHub repo creation").font = body_font()
ws_sum.cell(row=11, column=2, value="Pending").font = body_font(color="AA0000")
ws_sum.cell(row=11, column=3, value="Needed for CI auto-deploy; add AWS_ACCESS_KEY_ID + AWS_SECRET_ACCESS_KEY as repo secrets").font = body_font()
ws_sum.cell(row=12, column=1, value="Terraform remote state").font = body_font()
ws_sum.cell(row=12, column=2, value="Deferred").font = body_font(color=C_DEFER_TXT)
ws_sum.cell(row=12, column=3, value="S3 bucket already provisioned; run 'terraform init -migrate-state' when ready").font = body_font()
for r in [11, 12]:
    for c in range(1, 4):
        ws_sum.cell(row=r, column=c).border = border
        ws_sum.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws_sum.row_dimensions[r].height = 30

# ─── Credentials sheet ─────────────────────────────────────────────────────────
C_HAVE_BG  = "D6ECD9"   # green  — already set
C_HAVE_TXT = "1E4D2B"
C_NEED_BG  = "FFF3CD"   # amber  — need to get
C_NEED_TXT = "5A3E00"
C_OPT_BG   = "F0F0F0"   # grey   — optional / Phase 4
C_OPT_TXT  = "555555"

ws_cred = wb.create_sheet("Credentials & Secrets")
ws_cred.column_dimensions["A"].width = 8    # status chip
ws_cred.column_dimensions["B"].width = 30   # service
ws_cred.column_dimensions["C"].width = 32   # env var(s)
ws_cred.column_dimensions["D"].width = 22   # where stored
ws_cred.column_dimensions["E"].width = 55   # how to get it

ws_cred.merge_cells("A1:E1")
ch = ws_cred["A1"]
ch.value = "VirtualOffice AI — Credentials & Secrets Registry"
ch.font = Font(bold=True, color=C_WHITE, size=13)
ch.fill = hdr_fill(C_HDR_PHASE)
ch.alignment = Alignment(horizontal="center", vertical="center")
ws_cred.row_dimensions[1].height = 28

CRED_HEADERS = ["Status", "Service / Key", "Env Var(s)", "Stored In", "How to Get It"]
for i, h in enumerate(CRED_HEADERS, 1):
    c = ws_cred.cell(row=2, column=i, value=h)
    c.font = col_font()
    c.fill = hdr_fill(C_HDR_COL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws_cred.row_dimensions[2].height = 20

# status: "have" | "need" | "optional"
CREDS = [
    # ── ALREADY CONFIGURED ───────────────────────────────────────────────────
    ("have", "Anthropic (Claude API)",
     "ANTHROPIC_API_KEY",
     "AWS Secrets Manager\n/voai/staging/anthropic-api-key",
     "Already set in ECS task. Rotate at: console.anthropic.com → API Keys"),

    ("have", "AWS (ECS deploy + ECR)",
     "AWS_ACCESS_KEY_ID\nAWS_SECRET_ACCESS_KEY",
     "~/.aws/credentials\n[voai-staging]",
     "Already set locally. Add as GitHub repo secrets for CI:\nSettings → Secrets → Actions → New secret"),

    ("have", "PostgreSQL (RDS staging)",
     "DATABASE_URL\nMIGRATIONS_DATABASE_URL",
     "AWS Secrets Manager\n/voai/staging/database-url",
     "Already provisioned on RDS ap-south-1. URL format:\npostgresql://voai_app:PASSWORD@HOST:5432/voai"),

    ("have", "Neo4j AuraDB",
     "NEO4J_URI\nNEO4J_USER\nNEO4J_PASSWORD",
     "AWS Secrets Manager\n/voai/staging/neo4j-*",
     "Already provisioned at neo4j.io/cloud/aura-free.\nURI: neo4j+s://16c05d35.databases.neo4j.io"),

    ("have", "Redis (ElastiCache staging)",
     "REDIS_URL",
     "AWS Secrets Manager\n/voai/staging/redis-url",
     "Already provisioned. URL format:\nredis://HOST:6379"),

    # ── NEED TO GET ──────────────────────────────────────────────────────────
    ("have", "LiveKit Cloud",
     "LIVEKIT_URL\nLIVEKIT_API_KEY\nLIVEKIT_API_SECRET",
     "AWS Secrets Manager\n/voai/staging/livekit-*\n(ECS task def :10)",
     "✅ Configured. Project: mysquad-3z2qrqie.livekit.cloud\n"
     "Rotate at: cloud.livekit.io → Settings → Keys"),

    ("have", "Deepgram (Speech-to-Text)",
     "DEEPGRAM_API_KEY",
     "AWS Secrets Manager\n/voai/staging/deepgram-api-key\n(ECS task def :11)",
     "✅ Configured. Rotate at: console.deepgram.com → API Keys"),

    ("have", "ElevenLabs — API Key",
     "ELEVENLABS_API_KEY",
     "AWS Secrets Manager\n/voai/staging/elevenlabs-api-key\n(ECS task def :12)",
     "✅ API key configured. Rotate at: elevenlabs.io → Profile → API Key"),
    ("have", "ElevenLabs — Voice IDs",
     "VOICE_ID_SARAH\nVOICE_ID_PRIYA\nVOICE_ID_MARCUS",
     "ECS task definition\nenvironment vars\n(task def :13)",
     "✅ Configured.\nSarah: ecp3DWciuUyW7BYM7II1\nPriya: ZqvIIuD5aI9JFejebHiH\nMarcus: bfGb7JTLUnZebZRiFYyq"),

    ("need", "Stripe (Billing)",
     "STRIPE_SECRET_KEY\nSTRIPE_PRICE_STARTER\nSTRIPE_PRICE_GROWTH\nSTRIPE_PRICE_ENTERPRISE",
     "AWS Secrets Manager\n(not yet created)",
     "1. Go to dashboard.stripe.com → Sign up / log in\n"
     "2. Developers → API Keys → Copy Secret key (sk_live_... or sk_test_...)\n"
     "3. Products → Add product → Create 3 products with recurring prices\n"
     "   Copy each Price ID (price_xxx) for STARTER / GROWTH / ENTERPRISE\n"
     "4. Store secret key in Secrets Manager\n"
     "5. Add all 4 env vars to ECS task definition"),

    ("need", "Cal.com (Expert scheduling)",
     "CALCOM_API_KEY\nCALCOM_EVENT_TYPE_ID",
     "ECS env var\n(not yet created)",
     "1. Go to cal.com → Sign up / log in\n"
     "2. Settings → Developer → API Keys → Add new key\n"
     "3. Event Types → Create event type (e.g. 'Expert Session 30min')\n"
     "   Copy the numeric event type ID from the URL\n"
     "4. Add both env vars to ECS task definition environment"),

    ("have", "GitHub repo",
     "— (code pushed)",
     "github.com/manuani/mysquad\n(main branch)",
     "✅ Repo created and code pushed.\nRemote: https://github.com/manuani/mysquad.git"),
    ("need", "GitHub CI/CD secrets",
     "AWS_ACCESS_KEY_ID\nAWS_SECRET_ACCESS_KEY",
     "GitHub repo secrets\n(not yet added)",
     "Run in terminal after gh auth login:\n"
     "gh secret set AWS_ACCESS_KEY_ID --repo manuani/mysquad --body \"AKIAYS5M73Q4H5NMV7FS\"\n"
     "gh secret set AWS_SECRET_ACCESS_KEY --repo manuani/mysquad --body \"$(aws configure get aws_secret_access_key --profile voai-staging)\"\n"
     "Then every push to main auto-deploys to ECS staging."),

    # ── OPTIONAL / PHASE 4 ───────────────────────────────────────────────────
    ("optional", "WorkOS (SSO / AuthKit)",
     "WORKOS_API_KEY\nWORKOS_CLIENT_ID",
     "AWS Secrets Manager\n(Phase 4)",
     "1. workos.com → Sign up → Dashboard → API Keys\n"
     "2. Copy API Key + Client ID from your application\n"
     "3. Configure redirect URIs to match your staging domain\n"
     "4. Add both vars to ECS task definition (replaces dev-mode header auth)"),

    ("optional", "AWS S3 (Object Store)",
     "OBJECT_STORE_BUCKET\nOBJECT_STORE_ACCESS_KEY_ID\nOBJECT_STORE_SECRET_ACCESS_KEY\nOBJECT_STORE_REGION",
     "ECS env (bucket already set)\nSecrets Manager for keys",
     "Bucket already provisioned. Create a scoped IAM user for app access:\n"
     "IAM → Users → Create user → Attach S3 policy → Security credentials → Access key"),

    ("optional", "Stripe Connect (Expert payouts)",
     "STRIPE_CONNECT_CLIENT_ID",
     "AWS Secrets Manager\n(Phase 4)",
     "1. dashboard.stripe.com → Connect → Get started\n"
     "2. Platform profile → copy Client ID\n"
     "3. Implement OAuth flow for experts to onboard their bank accounts"),

    ("optional", "Admin API Key (internal)",
     "ADMIN_API_KEY",
     "ECS env var\n(default: dev-admin-key)",
     "Generate a strong random key:\n"
     "  openssl rand -hex 32\n"
     "Set as ADMIN_API_KEY in ECS task definition environment.\n"
     "Share only with ops team — protects all /v1/admin/* endpoints"),

    ("optional", "Scheduler Secret (internal)",
     "SCHEDULER_SECRET",
     "ECS env var\n(not yet set)",
     "Generate: openssl rand -hex 32\n"
     "Set same value in both scheduler service and api-server ECS envs.\n"
     "Used to authenticate x-scheduler-secret header on internal cron calls"),
]

CRED_STATUS_STYLE = {
    "have":     (C_HAVE_BG,  C_HAVE_TXT, "✅ Have"),
    "need":     (C_NEED_BG,  C_NEED_TXT, "🔴 Need"),
    "optional": (C_OPT_BG,   C_OPT_TXT,  "⚪ Optional"),
}

for r_idx, (status, service, envvars, stored_in, how_to) in enumerate(CREDS, 3):
    bg, fg, label = CRED_STATUS_STYLE[status]
    fill = row_fill(bg)
    vals = [label, service, envvars, stored_in, how_to]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_cred.cell(row=r_idx, column=c_idx, value=val)
        cell.font = body_font(color=fg, bold=(c_idx == 2))
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_cred.row_dimensions[r_idx].height = 80

# Freeze header
ws_cred.freeze_panes = "A3"

# Put Summary first
wb.active = ws_sum

out_path = "/Users/muralis/projects/mysquad/VoAI-Product-Tracker.xlsx"
wb.save(out_path)
print(f"Saved → {out_path}")
print(f"Totals: Done={counts['Done']}, Active={counts['Active']}, Pending={counts['Pending']}, Deferred={counts['Deferred']}, Total={total}")
